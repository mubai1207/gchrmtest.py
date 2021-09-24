#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import xlrd

__author__ = 'RaoPQ'

import os, sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from DemoGChrm.config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import configparser as cparser

# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG, encoding='UTF-8')
name = cf.get("tester", "name")


class WriteExcel:
    """写入excel文件"""

    def __init__(self, fileName, sheet_name=None):
        self.filename = fileName
        self.wb = load_workbook(fileName)
        self.data = xlrd.open_workbook(fileName)
        if sheet_name is None:
            self.table = self.data.sheets()[0]
            # self.ws = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])  # self.wb['%s' % sheet_name]
            self.ws = self.wb.active
        else:
            self.table = self.data.sheet_by_name(sheet_name)
            self.ws = self.wb['%s' % sheet_name]
            # self.ws = self.wb.get_sheet_by_name(sheet_name)
        # self.table = self.data.sheet_by_name(sheet_name)
        # self.n_rows = self.table.nrows # 获取总行数、
        self.n_cols = self.table.ncols  # 总列数

    def write_data(self, row_n, value):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """

        align = Alignment(horizontal='center', vertical='center')
        N_n = chr(self.n_cols - 1 + 64) + str(row_n)
        O_o = chr(self.n_cols + 64) + str(row_n)

        if value == "PASS":
            self.ws.cell(row_n, self.n_cols - 1, value)
            self.ws[N_n].font = Font(name='宋体', color='008000', bold=True)  # PASS颜色
        elif value == "FAIL":
            self.ws.cell(row_n, self.n_cols - 1, value)
            self.ws[N_n].font = Font(name='宋体', color='FF6100', bold=True)  # FAIL颜色
        elif value == "Error":
            self.ws.cell(row_n, self.n_cols - 1, value)
            self.ws[N_n].font = Font(name='宋体', color='FF0000', bold=True)  # Error颜色
        else:
            self.ws.cell(row_n, self.n_cols - 1, "NA")
            self.ws[N_n].font = Font(name='宋体', color='808080', bold=True)  # Skip颜色
        self.ws.cell(row_n, self.n_cols, name)
        self.ws[N_n].alignment = align
        self.ws[O_o].font = Font(name='宋体', color='0000FF', bold=True)
        self.ws[O_o].alignment = align
        self.wb.save(self.filename)
